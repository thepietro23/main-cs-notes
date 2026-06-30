# -*- coding: utf-8 -*-
"""Builds ML_System_Design_Master_Tracker.xlsx (matches DBMS/DSA tracker style)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ---- palette ----
NAVY   = "1F3864"   # title bar
BLUE   = "2E5496"   # header row
LBLUE  = "D6E4F0"   # band
GREEN  = "C6EFCE"
YELLOW = "FFEB9C"
GREY   = "F2F2F2"
WHITE  = "FFFFFF"

title_font  = Font(name="Calibri", size=14, bold=True, color=WHITE)
head_font   = Font(name="Calibri", size=11, bold=True, color=WHITE)
cell_font   = Font(name="Calibri", size=10, color="000000")
bold_font   = Font(name="Calibri", size=10, bold=True)
title_fill  = PatternFill("solid", fgColor=NAVY)
head_fill   = PatternFill("solid", fgColor=BLUE)
band_fill   = PatternFill("solid", fgColor=LBLUE)
grey_fill   = PatternFill("solid", fgColor=GREY)
wrap        = Alignment(wrap_text=True, vertical="top")
center      = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = border

def title_bar(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = title_font; c.fill = title_fill
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 30

def body(ws, r, ncols, band=False):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = cell_font; cell.alignment = wrap; cell.border = border
        if band: cell.fill = band_fill

# =====================================================================
# 00 README
# =====================================================================
ws = wb.active; ws.title = "00 README"
title_bar(ws, "ML SYSTEM DESIGN MASTERY  —  FAANG / AI-Engineer / Staff-Level Tracker", 6)
readme = [
    ("PURPOSE", "Life-long, reusable mastery tracker for ML System Design — interview prep (FAANG/AI labs) + AI/Backend/MLE roles + exam mapping (SEBI/RBI/GATE)."),
    ("PRIMARY TARGET", "ML System Design interviews at Google, Meta, Amazon, OpenAI, Anthropic, NVIDIA, Uber, Airbnb, Stripe + AI Engineering / MLOps roles."),
    ("HOW TO USE", "1) '01 Master Syllabus' = 23-module roadmap. 2) Follow '02 Study Plan' (12 weeks). 3) Drill '03 Interview Bank'. 4) Track in '06 Dashboard'."),
    ("NOTES FORMAT", "Each module = a rich .docx (Markdown -> Pandoc) with Mermaid/PNG diagrams, capacity numbers, trade-off tables, MCQs, mock interviews, cheat sheet."),
    ("MASTER PROMPT", "Use 00_MASTER_PROMPT_ML_System_Design.md to regenerate/extend any module with any LLM."),
    ("STATUS LEGEND", "Not Started  |  Drafting  |  Done  |  Reviewed  |  Revised"),
    ("PRIORITY LEGEND", "P0 = must-know core | P1 = important | P2 = good-to-know / advanced"),
]
r = 3
for k, v in readme:
    ws.cell(row=r, column=1, value=k).font = bold_font
    ws.cell(row=r, column=1).fill = grey_fill
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    c = ws.cell(row=r, column=2, value=v); c.font = cell_font; c.alignment = wrap
    ws.row_dimensions[r].height = 42
    r += 1
ws.column_dimensions["A"].width = 18
for col in "BCDEF": ws.column_dimensions[col].width = 22

# =====================================================================
# 01 Master Syllabus
# =====================================================================
ws = wb.create_sheet("01 Master Syllabus")
headers = ["Module", "Topic", "Subtopics", "Priority",
           "SEBI / RBI / GATE / Interview", "Difficulty", "Status", "Last Revised", "Notes File"]
title_bar(ws, "MASTER SYLLABUS — 23 Modules (beginner -> Staff-level)", len(headers))
for c, h in enumerate(headers, 1):
    ws.cell(row=2, column=c, value=h)
style_header(ws, 2, len(headers))

modules = [
 ("M01 Foundations of ML Systems", "What ML systems are; lifecycle; why they differ",
  "• Software 1.0 vs 2.0 • ML lifecycle • batch/online/streaming/on-device/federated • feedback loops & decay • ML vs classic system design",
  "P0", "SEBI★★ RBI★★ GATE★★ Interview★★★★", "Easy", "ML_SD_M01_Foundations.docx"),
 ("M02 ML SD Interview Framework", "The universal 7-step design framework",
  "• clarify reqs • frame ML problem • metrics (biz+ML, offline+online) • data/features • model • serving/scale • monitoring • jr vs staff signals",
  "P0", "Interview★★★★★ GATE★", "Easy", "ML_SD_M02_Framework.docx"),
 ("M03 Problem Framing & Requirements", "Business -> ML objective; task selection; sizing",
  "• biz goal -> ML objective • task choice (clf/reg/rank/gen/retrieval/RL) • when NOT to use ML • label design & leakage • capacity estimation (QPS/storage/latency)",
  "P0", "Interview★★★★★ GATE★★", "Easy-Med", "ML_SD_M03_Problem_Framing.docx"),
 ("M04 Data Engineering for ML", "Sources, pipelines, labeling, quality, versioning",
  "• batch vs streaming (Kafka/Kinesis) • orchestration (Airflow/Spark/Flink) • lake/warehouse/lakehouse • labeling & weak supervision • validation (TFDV/GE) • imbalance & sampling • DVC/lineage",
  "P0", "SEBI★★★ RBI★★★ GATE★★ Interview★★★★", "Med", "ML_SD_M04_Data_Engineering.docx"),
 ("M05 Feature Engineering & Stores", "Features, embeddings, training-serving skew, feature stores",
  "• feature types & encodings • hashing trick • embeddings • feature crosses • normalization • TRAINING-SERVING SKEW • feature stores (Feast/Tecton) • point-in-time correctness",
  "P0", "SEBI★★ RBI★★ GATE★★ Interview★★★★★", "Med", "ML_SD_M05_Features_Stores.docx"),
 ("M06 Model Development & Training", "Model choice, math of loss/optim, distributed training",
  "• linear->GBDT->DNN->transformer (when each) • baselines • bias-variance • loss fns & math • SGD/Adam • data/model/tensor/pipeline parallelism • ZeRO/FSDP • HPO • transfer learning",
  "P1", "SEBI★★ RBI★★ GATE★★★ Interview★★★★", "Med-Hard", "ML_SD_M06_Model_Training.docx"),
 ("M07 Model Evaluation", "Offline metrics by task; A/B testing; calibration",
  "• clf/reg/ranking/NLP/CV metrics • NDCG/MAP/MRR • ROC/PR-AUC • calibration • temporal splits & leakage • A/B testing & significance • interleaving • bandits • offline-online gap",
  "P0", "SEBI★★★ RBI★★ GATE★★★ Interview★★★★★", "Med", "ML_SD_M07_Evaluation.docx"),
 ("M08 Model Serving & Inference", "Batch vs online; latency; optimization; deployment",
  "• batch/online/streaming inference • serving patterns • p50/p95/p99 tail latency • TF Serving/Triton/KServe • quantization/pruning/distillation • CPU/GPU/TPU • caching & batching • shadow/canary",
  "P0", "GATE★ Interview★★★★★ Backend★★★★", "Med-Hard", "ML_SD_M08_Serving.docx"),
 ("M09 ML Infrastructure & MLOps", "MLOps maturity, registries, CI/CD/CT, orchestration",
  "• MLOps ladder • experiment tracking (MLflow/W&B) • model registry • CI/CD/CT • reproducibility • Kubeflow/Metaflow/TFX/SageMaker • k8s for ML • cost & GPU scheduling • Michelangelo/FBLearner",
  "P1", "Interview★★★★ Backend★★★★", "Med", "ML_SD_M09_MLOps.docx"),
 ("M10 Monitoring, Drift & Reliability", "What to monitor; drift detection; retraining",
  "• data/feature/prediction/concept drift • PSI/KL/KS math • retraining (scheduled/triggered/online) • feedback-loop dangers • shadow/canary/rollback/kill-switch • incident response",
  "P0", "Interview★★★★ Backend★★★", "Med", "ML_SD_M10_Monitoring_Drift.docx"),
 ("M11 Scaling ML Systems", "Scaling data/training/serving; sharding; cost triangle",
  "• scale data/train/serve independently • sharding/replication/partitioning • 1000s-GPU training • sharded embedding tables • autoscaling inference • cost/latency/throughput triangle",
  "P1", "GATE★★ Interview★★★★", "Hard", "ML_SD_M11_Scaling.docx"),
 ("M12 Recommendation Systems", "CF, MF, two-stage, two-tower, deep recommenders",
  "• CF (user/item) & MF math • content-based & hybrid • candidate-gen -> ranking • two-tower retrieval • Wide&Deep/DeepFM/DLRM/DCN • sequential (SASRec) • cold start • bandits • YouTube/TikTok/Netflix",
  "P0", "Interview★★★★★", "Hard", "ML_SD_M12_Recommendation.docx"),
 ("M13 Search & Ranking Systems", "IR foundations, Learning-to-Rank, neural retrieval",
  "• inverted index/TF-IDF/BM25 • LTR pointwise/pairwise/listwise (RankNet/LambdaMART) • dense/semantic retrieval + ANN • multi-stage ranking • query understanding • NDCG eval",
  "P0", "Interview★★★★★", "Hard", "ML_SD_M13_Search_Ranking.docx"),
 ("M14 Computer Vision Systems", "Classification/detection/segmentation; visual search",
  "• image clf/detection/segmentation pipelines • image embeddings & visual search • augmentation pipelines • serving CV (batching/GPU) • moderation/OCR/medical case studies",
  "P2", "GATE★★ Interview★★★", "Med-Hard", "ML_SD_M14_Computer_Vision.docx"),
 ("M15 NLP Systems", "Text pipelines, tasks, semantic similarity, serving",
  "• tokenization & embeddings (word2vec->BERT) • clf/NER/QA/summarization/translation • semantic search • serving latency • spam/abuse/autocomplete case studies",
  "P1", "GATE★★ Interview★★★", "Med-Hard", "ML_SD_M15_NLP.docx"),
 ("M16 LLM System Design", "RAG, vector DBs, fine-tuning, inference optimization, agents",
  "• tokens/context/cost • prompt vs RAG vs fine-tune • RAG end-to-end (chunk/embed/retrieve/rerank/gen) • hybrid search/HyDE • HNSW/IVF/PQ/FAISS • LoRA/QLoRA • KV cache/vLLM/paged-attn/spec-decoding/quant • agents & MCP • guardrails • LLM evals • cost routing",
  "P0", "Interview★★★★★ AI★★★★★", "Hard", "ML_SD_M16_LLM_System_Design.docx"),
 ("M17 Flagship Case Studies", "Full end-to-end designs with the 7-step framework",
  "• YouTube recs • FB/IG feed ranking • Ad CTR • fraud detection • Uber ETA • web search ranking • Airbnb search • spam/harmful content • e-commerce recs • PYMK • autocomplete • RAG assistant • multimodal",
  "P0", "Interview★★★★★", "Hard", "ML_SD_M17_Case_Studies.docx"),
 ("M18 Responsible & Trustworthy AI", "Fairness, privacy, explainability, security, governance",
  "• bias sources & metrics (demographic parity/equalized odds) • differential privacy & federated learning • SHAP/LIME • adversarial/poisoning/model-stealing • model cards • EU AI Act • LLM safety",
  "P1", "SEBI★★★ RBI★★★ GATE★ Interview★★★", "Med", "ML_SD_M18_Responsible_AI.docx"),
 ("M19 Systems Foundations for ML", "Distributed systems building blocks for ML",
  "• load balancing/caching (Redis)/CDN • SQL/NoSQL/KV/vector/time-series DBs • Kafka/Flink streaming • CAP & consistency • microservices/REST/gRPC • SLA/SLO • back-of-envelope estimation",
  "P0", "SEBI★★★ RBI★★★ GATE★★★ Interview★★★★★", "Med", "ML_SD_M19_Systems_Foundations.docx"),
 ("M20 Interview Mastery", "Question bank, 50+ mock designs, company styles",
  "• categorized question bank • fully worked mocks • Google/Meta/Amazon(LP)/Netflix/OpenAI/Anthropic/NVIDIA styles • handling curveballs • communication & whiteboarding • red flags",
  "P0", "Interview★★★★★", "Med", "ML_SD_M20_Interview_Mastery.docx"),
 ("M21 Hands-On Projects", "Build the systems end-to-end",
  "• recommender (candgen+rank) • semantic search w/ vector DB • RAG chatbot + evals • real-time fraud pipeline • mini feature store • A/B + monitoring dashboard • LLM serving w/ batching",
  "P1", "Interview★★★★ AI★★★★★", "Hard", "ML_SD_M21_Projects.docx"),
 ("M22 Revision & Roadmap", "Cheat sheets, flashcards, decision trees, mind maps",
  "• per-module cheat sheets • 60-sec framework card • metric-selection flowchart • decision trees (batch/online, RAG/fine-tune, GBDT/DNN) • flashcards • mind maps • 12-week + 2-week crash plan",
  "P0", "Interview★★★★★", "Easy", "ML_SD_M22_Revision_Roadmap.docx"),
 ("M23 Competitive Exam Mapping", "Map ML/data/systems topics to SEBI/RBI/GATE/ISRO",
  "• map to SEBI-IT/RBI-IT/GATE-CS&DA/ISRO/DRDO • highlight frequently-asked (ML basics, metrics, stats, DB for ML, pipelines) • PYQ-style Qs • flag interview-only vs exam-relevant",
  "P1", "SEBI★★★★ RBI★★★★ GATE★★★★", "Med", "ML_SD_M23_Exam_Mapping.docx"),
]
r = 3
for m in modules:
    mod, topic, subs, prio, rating, diff, fname = m
    vals = [mod, topic, subs, prio, rating, diff, "Not Started", "", fname]
    for c, v in enumerate(vals, 1):
        ws.cell(row=r, column=c, value=v)
    body(ws, r, len(headers), band=(r % 2 == 0))
    ws.cell(row=r, column=1).font = bold_font
    ws.cell(row=r, column=4).alignment = center
    ws.cell(row=r, column=6).alignment = center
    ws.cell(row=r, column=7).alignment = center
    ws.row_dimensions[r].height = 60
    r += 1
widths = [30, 30, 60, 8, 26, 11, 13, 13, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A3"

# =====================================================================
# 02 Study Plan (12 weeks)
# =====================================================================
ws = wb.create_sheet("02 Study Plan")
headers = ["Week", "Focus", "Modules", "Key Deliverable", "Status"]
title_bar(ws, "12-WEEK STUDY PLAN (+ 2-week interview crash variant in M22)", len(headers))
for c, h in enumerate(headers, 1): ws.cell(row=2, column=c, value=h)
style_header(ws, 2, len(headers))
plan = [
 ("1", "Foundations + the framework", "M01, M02", "Memorize the 7-step framework; explain ML vs classic SD"),
 ("2", "Problem framing + capacity estimation", "M03, M19", "Do 5 back-of-envelope sizings (QPS/storage/latency)"),
 ("3", "Data engineering + pipelines", "M04", "Design a batch+streaming data pipeline"),
 ("4", "Features + feature stores + skew", "M05", "Explain training-serving skew + design a feature store"),
 ("5", "Modeling + training at scale", "M06", "Pick models w/ justification; explain distributed training"),
 ("6", "Evaluation + A/B testing", "M07", "Choose metrics for 5 problems; design an A/B test"),
 ("7", "Serving + inference optimization", "M08", "Hit a latency budget; quantize/distill a model"),
 ("8", "MLOps + monitoring + drift", "M09, M10", "Design a retraining + drift-detection loop"),
 ("9", "Recommendation systems", "M12", "Design YouTube/e-commerce recs (candgen+rank)"),
 ("10", "Search & ranking + LLM systems", "M13, M16", "Design search ranking + a RAG assistant"),
 ("11", "Scaling + CV/NLP + Responsible AI", "M11, M14, M15, M18", "Scale a system 100x; add fairness/privacy"),
 ("12", "Case studies + mock interviews", "M17, M20, M21, M22, M23", "Do 5 full mock designs; build 1 project; exam mapping"),
]
r = 3
for row in plan:
    for c, v in enumerate(list(row) + ["Not Started"], 1):
        ws.cell(row=r, column=c, value=v)
    body(ws, r, len(headers), band=(r % 2 == 0))
    ws.cell(row=r, column=1).alignment = center
    ws.cell(row=r, column=5).alignment = center
    ws.row_dimensions[r].height = 34
    r += 1
for i, w in enumerate([8, 34, 22, 50, 14], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A3"

# =====================================================================
# 03 Interview Question Bank
# =====================================================================
ws = wb.create_sheet("03 Interview Bank")
headers = ["#", "Design Question", "Pattern / Core Skill", "Asked At", "Module(s)"]
title_bar(ws, "ML SYSTEM DESIGN INTERVIEW QUESTION BANK (high-frequency)", len(headers))
for c, h in enumerate(headers, 1): ws.cell(row=2, column=c, value=h)
style_header(ws, 2, len(headers))
bank = [
 ("Design YouTube / video recommendations", "Two-stage: candidate gen + ranking", "Google, Meta, Netflix", "M12, M17"),
 ("Design Facebook/Instagram news feed ranking", "Multi-objective ranking, freshness", "Meta, LinkedIn", "M12, M13, M17"),
 ("Design an ad click-through-rate (CTR) prediction system", "Calibrated probability, online learning", "Google, Meta, Amazon", "M07, M12, M17"),
 ("Design a fraud / anomaly detection system", "Imbalanced data, streaming features, low latency", "Stripe, Amazon, Uber", "M04, M07, M17"),
 ("Design ETA / delivery-time prediction", "Regression, spatial features, real-time", "Uber, DoorDash", "M05, M17"),
 ("Design web search ranking", "Retrieval + LTR, NDCG", "Google, Microsoft", "M13, M17"),
 ("Design Airbnb search / similar-listings", "Personalized ranking, embeddings", "Airbnb", "M13, M17"),
 ("Design a spam / harmful-content detection system", "Classification, adversarial drift", "Meta, Google", "M10, M15, M18"),
 ("Design a RAG-based enterprise knowledge assistant", "Chunk/embed/retrieve/rerank/generate + evals", "OpenAI, Anthropic, startups", "M16, M17"),
 ("Design a large-scale LLM serving system", "KV cache, batching, paged attention, GPU cost", "OpenAI, Anthropic, NVIDIA", "M08, M11, M16"),
 ("Design a semantic search engine over images/text", "Embeddings + ANN (HNSW/IVF)", "Pinterest, Google", "M13, M14, M16"),
 ("Design People-You-May-Know / friend recommendation", "Graph features, candidate gen", "LinkedIn, Meta", "M12, M17"),
 ("Design an autocomplete / typeahead system", "Trie + ranking, latency budget", "Google, Amazon", "M13, M17"),
 ("Design a recommendation cold-start strategy", "Content features, bandits, exploration", "Netflix, Spotify", "M12, M17"),
 ("Design a feature store / training-serving consistency layer", "Point-in-time correctness, online+offline", "Uber, DoorDash, Stripe", "M05, M09"),
 ("Design a model monitoring + auto-retraining pipeline", "Drift detection, triggers, rollback", "Most companies", "M09, M10"),
 ("Design a multimodal recommendation / search system", "Fused embeddings, multi-tower", "TikTok, Pinterest", "M12, M14, M16"),
 ("Design content moderation at scale", "CV+NLP, human-in-loop, precision/recall trade-off", "Meta, TikTok", "M14, M15, M18"),
]
r = 3
for i, (q, pat, at, mods) in enumerate(bank, 1):
    for c, v in enumerate([i, q, pat, at, mods], 1):
        ws.cell(row=r, column=c, value=v)
    body(ws, r, len(headers), band=(r % 2 == 0))
    ws.cell(row=r, column=1).alignment = center
    ws.row_dimensions[r].height = 30
    r += 1
for i, w in enumerate([5, 52, 38, 26, 16], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A3"

# =====================================================================
# 04 Exam Mapping
# =====================================================================
ws = wb.create_sheet("04 Exam Mapping")
headers = ["Topic Area", "SEBI IT", "RBI IT", "GATE CS/DA", "ISRO/DRDO/Other", "Note"]
title_bar(ws, "EXAM MAPPING — which ML/data/systems topics matter for which exam", len(headers))
for c, h in enumerate(headers, 1): ws.cell(row=2, column=c, value=h)
style_header(ws, 2, len(headers))
exam = [
 ("ML basics & lifecycle (M01)", "Medium", "Medium", "Med (DA)", "Medium", "Definitions, supervised/unsupervised"),
 ("Data engineering / pipelines (M04)", "Medium", "Medium", "Low-Med", "Low", "ETL, warehousing overlap with DBMS"),
 ("Feature engineering & stats (M05)", "Medium", "Medium", "High (DA)", "Medium", "Encoding, normalization, stats"),
 ("Model training & ML algorithms (M06)", "Low-Med", "Low-Med", "High (DA)", "Medium", "GATE-DA covers ML algos in depth"),
 ("Evaluation metrics (M07)", "Medium", "Medium", "High (DA)", "Medium", "Precision/recall/AUC frequently asked"),
 ("Systems foundations (M19)", "High", "High", "High (CS)", "High", "Caching, DBs, CAP, queues — strong overlap"),
 ("Responsible AI / privacy/security (M18)", "High", "High", "Low", "Medium", "SEBI/RBI care about governance & security"),
 ("ML System Design (M02,M08,M12,M16,M17)", "Low", "Low", "Low", "Low", "INTERVIEW-ONLY: rarely on written exams"),
 ("LLM/RAG/vector search (M16)", "Low", "Low", "Low-Med", "Low", "Emerging; mostly interview/AI-role relevant"),
]
r = 3
for row in exam:
    for c, v in enumerate(row, 1):
        ws.cell(row=r, column=c, value=v)
    body(ws, r, len(headers), band=(r % 2 == 0))
    for c in range(2, 6): ws.cell(row=r, column=c).alignment = center
    ws.row_dimensions[r].height = 30
    r += 1
for i, w in enumerate([40, 10, 10, 12, 18, 36], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A3"

# =====================================================================
# 05 Interview Focus (by role)
# =====================================================================
ws = wb.create_sheet("05 Interview Focus")
headers = ["Role", "ML SD Focus", "Must-Know", "Key Modules"]
title_bar(ws, "INTERVIEW / ROLE FOCUS — what each role emphasizes", len(headers))
for c, h in enumerate(headers, 1): ws.cell(row=2, column=c, value=h)
style_header(ws, 2, len(headers))
roles = [
 ("ML Engineer (MLE)", "End-to-end design: data->model->serve->monitor", "7-step framework, recsys/ranking, training-serving skew, A/B", "M02-M13, M17"),
 ("AI Engineer / LLM", "RAG, vector DBs, LLM serving, agents, evals", "RAG pipeline, HNSW/IVF, KV cache/vLLM, fine-tune vs prompt, guardrails", "M16, M17, M21"),
 ("ML Infra / Platform", "Training & serving infra, scaling, MLOps", "Distributed training, feature store, model registry, autoscaling, cost", "M06, M08, M09, M11"),
 ("Backend SDE (ML-adjacent)", "Serving infra, systems foundations", "Latency budgets, caching, queues, CAP, gRPC, capacity estimation", "M08, M19"),
 ("Data Engineer", "Pipelines, feature stores, data quality", "Batch vs streaming, Kafka/Flink, point-in-time correctness, validation", "M04, M05"),
 ("Applied Scientist", "Modeling depth + evaluation rigor", "Loss/metric math, calibration, experimental design, ranking losses", "M06, M07, M12, M13"),
]
r = 3
for row in roles:
    for c, v in enumerate(row, 1):
        ws.cell(row=r, column=c, value=v)
    body(ws, r, len(headers), band=(r % 2 == 0))
    ws.cell(row=r, column=1).font = bold_font
    ws.row_dimensions[r].height = 46
    r += 1
for i, w in enumerate([26, 42, 52, 22], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A3"

# =====================================================================
# 06 Progress Dashboard
# =====================================================================
ws = wb.create_sheet("06 Progress Dashboard")
title_bar(ws, "PROGRESS DASHBOARD", 3)
ws.cell(row=3, column=1, value="Module status (auto-counts from '01 Master Syllabus')").font = bold_font
statuses = ["Not Started", "Drafting", "Done", "Reviewed", "Revised"]
r = 4
for s in statuses:
    ws.cell(row=r, column=1, value=s).font = cell_font
    ws.cell(row=r, column=2, value=f"=COUNTIF('01 Master Syllabus'!G3:G25,\"{s}\")")
    ws.cell(row=r, column=1).border = border
    ws.cell(row=r, column=2).border = border
    ws.cell(row=r, column=2).alignment = center
    r += 1
ws.cell(row=r, column=1, value="TOTAL MODULES").font = bold_font
ws.cell(row=r, column=2, value="=COUNTA('01 Master Syllabus'!A3:A25)").alignment = center
ws.cell(row=r, column=1).border = border; ws.cell(row=r, column=2).border = border
r += 2
ws.cell(row=r, column=1, value="% Complete (Done+Reviewed+Revised)").font = bold_font
ws.cell(row=r, column=2,
        value="=ROUND((COUNTIF('01 Master Syllabus'!G3:G25,\"Done\")+COUNTIF('01 Master Syllabus'!G3:G25,\"Reviewed\")+COUNTIF('01 Master Syllabus'!G3:G25,\"Revised\"))/COUNTA('01 Master Syllabus'!A3:A25)*100,1)")
ws.cell(row=r, column=2).alignment = center
ws.cell(row=r, column=1).border = border; ws.cell(row=r, column=2).border = border
ws.column_dimensions["A"].width = 42
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 14

out = r"c:/AAA/Personal/DSA/ML_System_Design/ML_System_Design_Master_Tracker.xlsx"
wb.save(out)
print("SAVED:", out)
